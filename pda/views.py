from django.shortcuts import render, redirect
from datetime import datetime,timedelta,date
from django.db.models import Sum,Q
from django.shortcuts import get_object_or_404
from .models import Patio, Paquete,AlbaranDevolucion,Costes,LineaArticulo,TipoTarea,Trabajador,Articulo
from .forms import PatioForm,PaqueteForm,AlbaranForm,LineaArticulo,LineaArticuloForm,TrabajadorForm
from django.contrib import messages
from django.contrib.auth import authenticate, login
from django.views.decorators.csrf import csrf_exempt
from django.views.generic import TemplateView
from openpyxl import Workbook
from django.http import HttpResponse
from .utils import get_articulos_dict,get_tipos_tarea_dict,get_trabajadores_dict
from collections import defaultdict
from pda.models import Articulo, Delegacion1, Delegacion2, Delegacion3, Delegacion4
import pandas as pd
from openpyxl import load_workbook
from django.conf import settings
from io import BytesIO
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter




def iniciar_tarea(request):
    if request.method == 'POST':
        form = PatioForm(request.POST)
        # Creamos una instancia del formulario con los datos enviados
        if form.is_valid():
            tarea = form.save(commit=False)  # Crear objeto sin guardar
            tarea.fecha = datetime.now().date()  # Fecha actual
            tarea.horaInicio = datetime.now().time().replace(microsecond=0)  # Hora inicio actual
            tarea.save()  # Guardar tarea
            return redirect('crear_paquete', tarea_id=tarea.id)  # Redirigir siguiente paso
    else:
        form = PatioForm()  # Mostrar formulario vacío

    trabajadores_dict = get_trabajadores_dict()  # Datos para formulario
    tipos_tarea_dict = get_tipos_tarea_dict()
 # Renderizamos la plantilla con el formulario y las listas de datos
    return render(request, 'iniciar_tarea.html', {
        'form': form,
        'trabajadores': trabajadores_dict,
        'tipos_tarea': tipos_tarea_dict,
    })


def crear_paquete(request, tarea_id):
    tarea = get_object_or_404(Patio, id=tarea_id)

    if request.method == 'POST':
        form = PaqueteForm(request.POST)
        if form.is_valid():
            ahora = datetime.now().time().replace(microsecond=0)
            
            # Si es el primer paquete, usamos la hora de inicio de la tarea
            if not Paquete.objects.filter(tarea=tarea).exists():
                hora_inicio_paquete = tarea.horaInicio  # Para el primer paquete, la hora es la de la tarea
                hora_fin_paquete = ahora  # Hora de fin del primer paquete, la hora actual
            else:
                # Para los paquetes siguientes, la hora de inicio es la hora de fin del último paquete
                ultimo_paquete = Paquete.objects.filter(tarea=tarea).order_by('-id').first()
                hora_inicio_paquete = ultimo_paquete.horaFin  # La hora de inicio será la hora de fin del último paquete
                hora_fin_paquete = ahora  # La hora de fin será la hora actual

            # Crea el paquete sin guardarlo todavía
            paquete = form.save(commit=False)
            paquete.tarea = tarea
            paquete.horaInicio = hora_inicio_paquete  # Asignamos la hora de inicio al paquete
            paquete.horaFin = hora_fin_paquete  # Asignamos la hora de fin al paquete

            # Guardamos el paquete
            paquete.save()
            print(f"Nuevo paquete creado: horaInicio = {paquete.horaInicio}, horaFin = {paquete.horaFin}")

            # Si ya existía un paquete, actualizamos la hora de fin del paquete anterior
            if Paquete.objects.filter(tarea=tarea).count() > 1:
                ultimo_paquete = Paquete.objects.filter(tarea=tarea).order_by('-id').first()
                ultimo_paquete.horaFin = hora_fin_paquete  # El último paquete termina cuando empieza el nuevo
                ultimo_paquete.save()
                print(f"Hora fin del paquete anterior (ID {ultimo_paquete.id}) actualizada a {hora_fin_paquete}")

            # Actualizar la cantidad total
            # aggregate(Sum(...)) devuelve un diccionario con la suma bajo la clave 'cantidad_paquete__sum'.
# Usamos ['cantidad_paquete__sum'] para acceder al valor, y 'or 0' por si no hay paquetes
            total_cantidad = Paquete.objects.filter(tarea=tarea).aggregate(Sum('cantidad_paquete'))['cantidad_paquete__sum'] or 0
            tarea.cantidad = total_cantidad
            tarea.save()

            # Redirige a la misma página de crear paquete o a donde desees
            return redirect('crear_paquete', tarea_id=tarea.id)

        else:
            # Si el formulario no es válido, asegúrate de devolver una respuesta
            # En este caso, simplemente renderizamos el formulario con los errores
            articulos_dict = get_articulos_dict()
            return render(request, 'crear_paquete.html', {
                'form': form,
                'tarea': tarea,
                'articulos': articulos_dict,
            })
    else:
        form = PaqueteForm()
        articulos_dict = get_articulos_dict()
        return render(request, 'crear_paquete.html', {
            'form': form,
            'tarea': tarea,
            'articulos': articulos_dict,
        })


def finalizar_tarea(request, tarea_id):
    tarea = get_object_or_404(Patio, id=tarea_id)
    tarea.horaFin = datetime.now().time().replace(microsecond=0)  
    tarea.save()
    return redirect('iniciar_tarea') 


def seleccionar_albaran(request):
    if request.method == 'POST':
        form = AlbaranForm(request.POST)  # Se instancia el formulario con los datos enviados
        if form.is_valid():
            numero = form.cleaned_data['numero']  # Se obtiene el número de albarán del formulario

            # Busca un albarán con ese número. Si no existe, lo crea.
            albaran, creado = AlbaranDevolucion.objects.get_or_create(numero=numero)

            return redirect('agregar_lineas', albaran_id=albaran.id)
        else:
            print("Este formulario ya ha sido creado:", form.errors)  # (opcional) muestra errores en consola
    else:
        form = AlbaranForm()  # Si es GET, muestra el formulario vacío

    return render(request, 'seleccionar_albaran.html', {'form': form})


class HomeView (TemplateView): 
    template_name='index.html'


def agregar_lineas(request, albaran_id):
    albaran = AlbaranDevolucion.objects.get(id=albaran_id)  # Se obtiene el albarán al que se le van a añadir líneas

    if request.method == 'POST':
        form = LineaArticuloForm(request.POST)  # Se instancia el formulario con los datos enviados

        if form.is_valid():
            # Se extraen los datos limpios del formulario
            cantidad_buena = form.cleaned_data['cantidad_buena']
            cantidad_mala = form.cleaned_data['cantidad_mala']
            chatarra = form.cleaned_data['chatarra']
            idArticulo = form.cleaned_data['idArticulo']

            # Se crea una nueva línea de artículo asociada al albarán
            LineaArticulo.objects.create(
                albaran=albaran,
                idArticulo=idArticulo,
                cantidad_buena=cantidad_buena,
                chatarra=chatarra,
                cantidad_mala=cantidad_mala
            )

            # Redirige a la misma página para seguir añadiendo más líneas
            return redirect('agregar_lineas', albaran_id=albaran.id)

        else:
            # Si el formulario es inválido, vuelve a mostrar el formulario con errores
            return render(request, 'agregar_lineas.html', {
                'form': form,
                'albaran': albaran,
                'articulos': get_articulos_dict(),
            })
    
    else:
        # Si es GET, muestra el formulario vacío
        form = LineaArticuloForm()
        articulos_dict = get_articulos_dict()
        return render(request, 'agregar_lineas.html', {
            'form': form, 
            'albaran': albaran,
            'articulos': articulos_dict
        })


def agregar_lineas2(request,albaran_id):
    albaran = AlbaranDevolucion.objects.get(id=albaran_id)
    if request.method=='POST':
        form=LineaArticuloForm(request.POST)
        if form.is_valid():
         cantidad_buena= form.cleaned_data['cantidad_buena']
         cantidad_mala= form.cleaned_data['cantidad_mala']
         chatarra= form.cleaned_data['chatarra']
         idArticulo= form.cleaned_data['idArticulo']

         LineaArticulo.objects.create(
             albaran=albaran,
             idArticulo=idArticulo,
             cantidad_buena=cantidad_buena,
             chatarra=chatarra,
             cantidad_mala=cantidad_mala
         )
         return redirect('añadir_lineas', albaran_id=albaran.id)
        else:           
            return render(request,'añadir_linea.html',{
                'form':form,
                'albaran':albaran,
                'articulos':get_articulos_dict(),
            })
    else:
        form = LineaArticuloForm()
        articulos_dict = get_articulos_dict()
        return render(request, 'añadir_linea.html', {
            'form': form, 
            'albaran': albaran,
            'articulos':articulos_dict
             })


def salir(request):
    return render(request,'cerrar_programa.html')


def estadisticas(request):

    return render(request,'estadisticas.html')


def lista_trabajadores(request):
    trabajadores = Trabajador.objects.all()
    return render(request, 'lista_trabajadores.html', {'trabajadores_lista': trabajadores})


def lista_tareas_completa(request):
    tareas = Patio.objects.all().order_by('-fecha')
    
    tareas_info = []
    for tarea in tareas:
        tarea_info = {
            'id': tarea.id,
            'fecha': tarea.fecha,
            'horaInicio': tarea.horaInicio,
            'horaFin': tarea.horaFin,
            'tipo_tarea': TipoTarea.objects.filter(id=tarea.idTipTarea).first(),
            'operador1': Trabajador.objects.filter(id=tarea.idOper1).first(),
            'operador2': Trabajador.objects.filter(id=tarea.idOper2).first() if tarea.idOper2 else None,
            'cantidad': tarea.cantidad
        }
        tareas_info.append(tarea_info)

    return render(request, 'lista_tareas.html', {
        'tareas_info': tareas_info
    })


def detalles_tarea(request,tarea_id):
    tarea=get_object_or_404(Patio,id=tarea_id)
    paquetes=Paquete.objects.filter(tarea=tarea)
    volver_url = request.META.get('HTTP_REFERER', '/lista_tareas/')
    return render(request,'detalle_tarea.html',{
        'tarea':tarea,
        'paquetes':paquetes,
        'volver_url':volver_url
        })


def estadisticas_trabajador(request, trabajador_id):
    trabajador = get_object_or_404(Trabajador, id=trabajador_id)
    tareas = Patio.objects.filter(Q(idOper1=trabajador_id) | Q(idOper2=trabajador_id))
    
    # Obtener el periodo de la query string (día, semana, mes, todo)
    periodo = request.GET.get('periodo')
    hoy = date.today()
    
    # Variables generales
    cantidadTotal = 0
    tiempoTotalSegundos = 0
    productividad = 0
    numTareas = 0
    tiempoPromedio = "00:00:00"
    
    # Diccionario para almacenar estadísticas por tipo de tarea
    estadisticas_por_tipo = defaultdict(lambda: {
        'cantidad': 0,
        'tiempo_total': 0,
        'num_tareas': 0,
        'tareas': []
    })
    
    # Filtrar tareas según el periodo
    if periodo == 'dia':
        tareas = tareas.filter(fecha=hoy)
    elif periodo == 'semana':
        inicio_semana = hoy - timedelta(days=hoy.weekday())  # Primer día de la semana
        tareas = tareas.filter(fecha__gte=inicio_semana, fecha__lte=hoy)
    elif periodo == 'mes':
        tareas = tareas.filter(fecha__month=hoy.month, fecha__year=hoy.year)
    elif periodo == 'todo':
        pass  # No filtra, usa todas las tareas

    # Agrupar las tareas por tipo
    for tarea in tareas:
        if tarea.horaInicio and tarea.horaFin and tarea.idTipTarea is not None:
            try:
                tipo_tarea = TipoTarea.objects.get(id=tarea.idTipTarea)
                nombre_tarea = tipo_tarea.nombre
            except TipoTarea.DoesNotExist:
                nombre_tarea = f"Tarea ID {tarea.idTipTarea}"

            # Calcular tiempo de la tarea
            fecha = tarea.fecha
            hora_inicio = datetime.combine(fecha, tarea.horaInicio)
            hora_fin = datetime.combine(fecha, tarea.horaFin)
            tiempo_segundos = (hora_fin - hora_inicio).total_seconds()

            # Acumular estadísticas por tipo de tarea
            estadisticas_por_tipo[nombre_tarea]['cantidad'] += tarea.cantidad or 0
            estadisticas_por_tipo[nombre_tarea]['tiempo_total'] += tiempo_segundos
            estadisticas_por_tipo[nombre_tarea]['num_tareas'] += 1
            estadisticas_por_tipo[nombre_tarea]['tareas'].append(tarea)

    # Calcular productividad y tiempo promedio
    for nombre_tarea, datos in estadisticas_por_tipo.items():
        tiempo_total = datos['tiempo_total']
        num_tareas = datos['num_tareas']
        cantidad = datos['cantidad']

        # Calcular productividad (cantidad por hora)
        if tiempo_total > 0:
            datos['productividad'] = round(cantidad / (tiempo_total / 3600), 2)
        else:
            datos['productividad'] = 0

        # Calcular tiempo promedio por tarea
        if num_tareas > 0:
            tiempo_promedio = tiempo_total / num_tareas
            h = int(tiempo_promedio // 3600)
            m = int((tiempo_promedio % 3600) // 60)
            s = int(tiempo_promedio % 60)
            datos['tiempo_promedio'] = f"{h:02d}:{m:02d}:{s:02d}"
        else:
            datos['tiempo_promedio'] = "00:00:00"

    # Preparar los datos para el template
    return render(request, 'estadisticas_trabajador.html', {
        'trabajador': trabajador,
        'tareas': tareas,
        'estadisticas_por_tipo': dict(estadisticas_por_tipo),
        'periodo': periodo,
    })

def lista_albaranes_completa(request):
    listaAlbaranes = AlbaranDevolucion.objects.all().order_by('-fecha')  # Obtiene todos los albaranes, ordenados por fecha descendente
    albaranesinfo = []

    for albaran in listaAlbaranes:
        # Se crea un diccionario con los datos reDelegacion2s de cada albarán
        albaraninfo = {
            'id': albaran.id,
            'numero': albaran.numero,
            'fecha': albaran.fecha
        }
        albaranesinfo.append(albaraninfo)

    # Se pasa la lista de albaranes a la plantilla
    return render(request, 'lista_albaranes.html', {
        'albaranesinfo': albaranesinfo
    })


def detalles_albaran(request, albaran_id):
    albaran = get_object_or_404(AlbaranDevolucion, id=albaran_id)  # Se obtiene el albarán o se muestra 404 si no existe
    articulos = LineaArticulo.objects.filter(albaran=albaran)  # Se obtienen todas las líneas del albarán

    articulos_dict = get_articulos_dict() 
    for linea in articulos:
        # Añade el nombre del artículo a cada línea
        linea.nombre_articulo = articulos_dict.get(linea.idArticulo)

    # Renderiza los detalles del albarán y sus líneas
    return render(request, 'detalle_albaran.html', {
        'albaran': albaran,
        'articulos': articulos,
    })



def eliminar_alabarn(request, albaran_id):
    albaran = get_object_or_404(AlbaranDevolucion, id=albaran_id)
    albaran.delete()
    return redirect('listaAlbaranes')


def eliminar_tarea(request,tarea_id):
    tarea=get_object_or_404(Patio,id=tarea_id)
    tarea.delete()
    return redirect('listaTareas')

@csrf_exempt 
def login_view(request):
    if request.method == 'POST':
        username = request.POST['username']
        password = request.POST['password']
        user = authenticate(request, username=username, password=password)

        if user is not None:
            login(request, user)
            print('Login correcto')
            return redirect('home')  # Esto depende de que la URL 'index' esté definida en tu URLconf
        else:
            print('Login fallido')
            messages.error(request, 'Usuario o contraseña incorrecta')

    return render(request, 'login.html')
 
    
def editar_linea_articulo(request, linea_id):
    linea = get_object_or_404(LineaArticulo, id=linea_id)
    albaran = linea.albaran

    if request.method == 'POST':
        form = LineaArticuloForm(request.POST, instance=linea)
        if form.is_valid():
            form.save()
            return redirect('detalle_albaran', albaran_id=albaran.id)
    else:
        form = LineaArticuloForm(instance=linea)

    return render(request, 'editar_linea.html', {
        'form': form,
        'albaran': albaran
    })


def eliminar_linea_articulo(request, linea_id):
    linea = get_object_or_404(LineaArticulo, id=linea_id)
    albaran_id = linea.albaran.id
    linea.delete()
    return redirect('detalle_albaran', albaran_id=albaran_id)


def exportar_trabajadores_excel(request):
   
   wb=Workbook()
   ws=wb.active
   ws.title='Trabajadores'
   ws.append(['Id','Nombre'])

   for trabajador in Trabajador.objects.all():
       ws.append([trabajador.id,trabajador.nombre])


   response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
   response['Content-Disposition'] = 'attachment; filename=trabajadores.xlsx'
   wb.save(response)
   return response


def comparativa_productividad(request):
    tareas_disponibles = TipoTarea.objects.all()
    periodo = request.GET.get('periodo')
    hoy = date.today()
    tareas = Patio.objects.all()
    
    # Filtros de fechas
    fecha_inicio_str = request.GET.get('fecha_inicio')
    fecha_fin_str = request.GET.get('fecha_fin')
    
    if fecha_inicio_str and fecha_fin_str:
        try:
            fecha_inicio = datetime.strptime(fecha_inicio_str, '%Y-%m-%d').date()
            fecha_fin = datetime.strptime(fecha_fin_str, '%Y-%m-%d').date()
            tareas = tareas.filter(fecha__range=(fecha_inicio, fecha_fin))
        except ValueError:
            pass
    
    # Filtros de tipo de tarea
    tarea_id = request.GET.get('tarea')
    if tarea_id:
        tareas = tareas.filter(idTipTarea=tarea_id)
    
   
    # Diccionario para estadísticas
    estadisticas_por_trabajador = defaultdict(lambda: {
        'cantidad': 0,
        'tiempo_total': 0,
        'num_tareas': 0,
        'productividad': 0,
        'nombre': '',
        'apellidos': '',
    })

    # Obtener tareas según el periodo
    for tarea in tareas:
        if tarea.horaInicio and tarea.horaFin:
            hora_inicio = datetime.combine(tarea.fecha, tarea.horaInicio)
            hora_fin = datetime.combine(tarea.fecha, tarea.horaFin)
            tiempo_segundos = (hora_fin - hora_inicio).total_seconds()
            
            print('tipo horaInicio,:',type(tarea.horaInicio))
            print('tipo horaFin: ',type(tarea.horaFin))

            # Obtener los trabajadores por su ID
            for operador_id in [tarea.idOper1, tarea.idOper2]:
                if operador_id:
                    # Buscar el trabajador correspondiente
                    try:
                        trabajador_obj = Trabajador.objects.get(id=operador_id)
                        stats = estadisticas_por_trabajador[operador_id]
                        stats['nombre'] = trabajador_obj.nombre
                        stats['apellidos'] = trabajador_obj.apellidos
                        stats['cantidad'] += tarea.cantidad or 0
                        stats['tiempo_total'] += tiempo_segundos
                        stats['num_tareas'] += 1
                    except Trabajador.DoesNotExist:
                        stats = estadisticas_por_trabajador[operador_id]
                        stats['nombre'] = f'Operador ID {operador_id} no encontrado'

    # Calcular productividad
    for stats in estadisticas_por_trabajador.values():
        if stats['tiempo_total'] > 0:
           stats['productividad'] = round(stats['cantidad'] / (stats['tiempo_total'] / 3600), 2)
           
           horas= int(stats['tiempo_total'] // 3600)
           minutos= int((stats['tiempo_total'] % 3600) // 60)
           segundos=int(stats['tiempo_total'] % 60)
           stats['tiempo_total_formateado']=f'{horas:02d}:{minutos:02d}:{segundos:02d}'
        else:
            stats['tiempo_total_formateado']= '00:00:00'

    # Ordenar trabajadores por productividad
    trabajadores_ordenados = sorted(estadisticas_por_trabajador.items(), key=lambda x: x[1]['productividad'], reverse=True)
    
      
    # Filtro por rango de productividad (nuevo)
    min_prod = request.GET.get('min_prod')
    max_prod = request.GET.get('max_prod')

    if min_prod:
        trabajadores_ordenados=[t for t in trabajadores_ordenados if t[1]['productividad'] >= float(min_prod)]
        
    if max_prod:
        trabajadores_ordenados=[t for t in trabajadores_ordenados if t[1]['productividad'] <= float(max_prod)]

    # Enviar los datos al template
    return render(request, 'comparativa_productividad.html', {
        'trabajadores': trabajadores_ordenados,
        'periodo': periodo,
        'tareas_disponibles': tareas_disponibles
    })


def exportar_datos(request):
    # Definimos las delegaciones y sus respectivos modelos
    delegaciones = ['Delegacion1', 'Delegacion2', 'Delegacion3', 'Delegacion4']
    modelos = {
        'Delegacion1': Delegacion1,
        'Delegacion2': Delegacion2,
        'Delegacion3': Delegacion3,
        'Delegacion4': Delegacion4
    }

    # Definimos las columnas del Excel
    columnas = ['Articulo', 'Nombre']
    subcolumnas = ['Tot.Fact.Alq.Dia', 'Tot.Unid', 'P.Alq.Medio', '%Fact']
    for _ in delegaciones + ['General']:
        columnas.extend(subcolumnas)
    columnas.extend(['Coste Ud.', 'Coste Total', 'Coste Delegacion1', 'Coste Delegacion3', 'Coste Delegacion4', 'Coste Delegacion2'])

    # Obtenemos los costes unitarios por artículo
    coste_por_articulo = {coste.articulo.id: coste.precio for coste in Costes.objects.all()}

    # Calculamos la facturación total por delegación
    total_fact_deleg = {}
    for deleg in delegaciones:
        total = 0
        for data in modelos[deleg].objects.all():
            total += data.tot_unid * data.p_alq_medio
        total_fact_deleg[deleg] = total / 100  # Convertimos a formato monetario

    # Preparación de los datos de artículos
    totales_por_deleg = {deleg: 0 for deleg in delegaciones}
    datos_articulos = []

    for articulo in Articulo.objects.all():
        # Si el artículo no está presente en ninguna delegación, lo omitimos
        if not any(modelos[deleg].objects.filter(articulo=articulo).exists() for deleg in delegaciones):
            continue

        fila = [articulo.id, articulo.nombre]
        general_total_fact = 0
        general_total_unid = 0
        data_por_deleg = {}

        # Recogemos la información de cada delegación
        for deleg in delegaciones:
            data = modelos[deleg].objects.filter(articulo=articulo).first()
            data_por_deleg[deleg] = data
            if data:
                tot = data.tot_unid * data.p_alq_medio / 100
                totales_por_deleg[deleg] += tot
                general_total_fact += tot
                general_total_unid += data.tot_unid

        # Guardamos los datos procesados para este artículo
        datos_articulos.append({
            'articulo': articulo,
            'fila_base': fila,
            'data_por_deleg': data_por_deleg,
            'general_total_fact': general_total_fact,
            'general_total_unid': general_total_unid
        })

    # Total general de facturación
    total_general = sum(totales_por_deleg.values())

    # Ordenamos por facturación general
    datos_articulos.sort(key=lambda x: x['general_total_fact'], reverse=True)

    # Creamos la lista de resultados final
    resultados = []
    for info in datos_articulos:
        fila = info['fila_base']
        general_total_fact = info['general_total_fact']
        general_total_unid = info['general_total_unid']
        data_por_deleg = info['data_por_deleg']

        # Añadimos los datos por cada delegación
        for deleg in delegaciones:
            data = data_por_deleg[deleg]
            if data:
                tot = data.tot_unid * data.p_alq_medio / 100
                p_medio = data.p_alq_medio / 100
                total_deleg = total_fact_deleg[deleg]
                porcentaje = (tot / total_deleg * 100) if total_deleg else 0
                fila += [f'{tot:,.2f}', f'{data.tot_unid:,}', f'{p_medio:.4f}', f'{porcentaje:.2f}%']
            else:
                fila += ['-', '-', '-', '-']

        # Datos generales
        if general_total_unid:
            p_general_medio = general_total_fact / general_total_unid
        else:
            p_general_medio = 0
        porcentaje_general = (general_total_fact / total_general * 100) if total_general else 0
        fila += [
            f'{general_total_fact:,.2f}', f'{general_total_unid:,}',
            f'{p_general_medio:.4f}', f'{porcentaje_general:.3f}%'
        ]

        # Costes por artículo
        articulo_id = info['articulo'].id
        coste_unitario = coste_por_articulo.get(articulo_id, 0)
        fila.append(f"{coste_unitario:,.4f}")  # Coste Ud.
        coste_total_general = coste_unitario * general_total_unid
        fila.append(f"{coste_total_general:,.2f}")

        # Coste por cada delegación
        for deleg in delegaciones:
            data = data_por_deleg[deleg]
            if data:
                coste_deleg = coste_unitario * data.tot_unid
                fila.append(f"{coste_deleg:,.2f}")
            else:
                fila.append('-')

        resultados.append(fila)

    # Rellenamos las filas incompletas
    for fila in resultados:
        while len(fila) < len(columnas):
            fila.append('-')

    # Creamos el DataFrame
    df = pd.DataFrame(resultados, columns=columnas)

    # Exportamos a Excel
    output = BytesIO()
    df.to_excel(output, index=False, header=False, startrow=4)
    output.seek(0)
    wb = load_workbook(output)
    ws = wb.active

    # Aplicamos estilos
    bold = Font(bold=True)
    center = Alignment(horizontal='center', vertical='center')
    border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    colores = {'Delegacion1': 'FFF2CC', 'Delegacion2': 'D9EAD3', 'Delegacion3': 'CFE2F3', 'Delegacion4': 'F4CCCC'}

    # Título
    fecha = datetime.today().strftime('%d/%m/%Y')
    titulo = f"Ranking Articulos Comparativo por Delegación a Fecha: {fecha}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(columnas))
    c = ws.cell(row=1, column=1)
    c.value = titulo
    c.font = Font(size=14, bold=True, color='FFFFFF')
    c.alignment = center
    c.fill = PatternFill(start_color='4F81BD', end_color='4F81BD', fill_type='solid')

    # Encabezados
    ws.merge_cells(start_row=2, start_column=1, end_row=3, end_column=1)
    ws.cell(row=2, column=1, value='Articulo').font = bold
    ws.cell(row=2, column=1).alignment = center

    ws.merge_cells(start_row=2, start_column=2, end_row=3, end_column=2)
    ws.cell(row=2, column=2, value='Nombre').font = bold
    ws.cell(row=2, column=2).alignment = center

    start_col = 3
    for deleg in delegaciones:
        ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + 3)
        cell = ws.cell(row=2, column=start_col)
        cell.value = deleg
        cell.font = bold
        cell.alignment = center
        fill = PatternFill(start_color=colores[deleg], end_color=colores[deleg], fill_type='solid')
        for col in range(start_col, start_col + 4):
            ws.cell(row=2, column=col).fill = fill
            ws.cell(row=3, column=col).fill = fill
        start_col += 4

    # General
    ws.merge_cells(start_row=2, start_column=start_col, end_row=2, end_column=start_col + 3)
    cell = ws.cell(row=2, column=start_col)
    cell.value = 'General'
    cell.font = bold
    cell.alignment = center
    for col in range(start_col, start_col + 4):
        ws.cell(row=3, column=col).fill = PatternFill(start_color='D9D9D9', end_color='D9D9D9', fill_type='solid')

    subcols = ['Tot.Fact.Alq.Dia', 'Tot.Unid', 'P.Alq.Medio', '%Fact']
    col = 3
    for _ in delegaciones + ['General']:
        for sub in subcols:
            cell = ws.cell(row=3, column=col)
            cell.value = sub
            cell.font = bold
            cell.alignment = center
            cell.border = border
            col += 1

    headers_coste = ['Coste Ud.', 'Coste Total', 'Coste Delegacion1', 'Coste Delegacion2', 'Coste Delegacion3', 'Coste Delegacion4']
    for i, h in enumerate(headers_coste):
        cell = ws.cell(row=3, column=col + i, value=h)
        cell.font = bold
        cell.alignment = center
        cell.border = border
        cell.fill = PatternFill(start_color='BDD7EE', end_color='BDD7EE', fill_type='solid')

    # Aplicamos bordes y alineación a las filas de datos
    for row_idx in range(5, 5 + len(resultados)):
        for col_idx in range(1, len(columnas) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center')

    # Ajustar anchos de columna para mejor visualización
    ws.column_dimensions['A'].width = 12  # Articulo (ID)
    ws.column_dimensions['B'].width = 45  # Nombre

    total_cols = len(columnas)
    for col_idx in range(3, total_cols + 1):
        col_letter = ws.cell(row=4, column=col_idx).column_letter
        ws.column_dimensions[col_letter].width = 15

    # --- SUMAS POR COLUMNA ---

    # Calculamos las sumas necesarias
    sum_tot_fact_alq_dia = {deleg: 0 for deleg in delegaciones + ['General']}
    sum_coste_total = 0

    for fila in resultados:
        # Sumas de Tot.Fact.Alq.Dia (columna 3,7,11,15,19 etc.)
        for i, deleg in enumerate(delegaciones):
            col_idx = 2 + i * 4  # índice Python (columna Excel -1)
            valor = fila[col_idx]
            if valor != '-' and valor != '':
                try:
                    sum_tot_fact_alq_dia[deleg] += float(valor.replace(',', ''))
                except:
                    pass
        # General Tot.Fact.Alq.Dia (columna después de delegaciones)
        col_idx_general = 2 + len(delegaciones) * 4
        valor = fila[col_idx_general]
        if valor != '-' and valor != '':
            try:
                sum_tot_fact_alq_dia['General'] += float(valor.replace(',', ''))
            except:
                pass
        # Coste Total (columna nombre que buscas)
        try:
            col_coste_total = columnas.index('Coste Total')
            valor = fila[col_coste_total]
            if valor != '-' and valor != '':
                sum_coste_total += float(valor.replace(',', ''))
        except:
            pass

    # Preparamos fila suma vacía
    fila_suma = [''] * len(columnas)
    fila_suma[0] = 'TOTAL'

    # Ponemos las sumas en su columna correspondiente (Tot.Fact.Alq.Dia)
    for i, deleg in enumerate(delegaciones):
        col_idx = 2 + i * 4
        fila_suma[col_idx] = f"{sum_tot_fact_alq_dia[deleg]:,.2f}"

    # General Tot.Fact.Alq.Dia
    col_idx_general = 2 + len(delegaciones) * 4
    fila_suma[col_idx_general] = f"{sum_tot_fact_alq_dia['General']:,.2f}"

    # Coste Total
    col_coste_total = columnas.index('Coste Total')
    fila_suma[col_coste_total] = f"{sum_coste_total:,.2f}"

    # Insertamos la fila suma justo después de los datos (fila 5 + len(resultados))
    fila_suma_row = 5 + len(resultados)
    for col_num, valor in enumerate(fila_suma, start=1):
        ws.cell(row=fila_suma_row, column=col_num, value=valor)

    # Aplicamos estilos a la fila de suma
    for col_num in range(1, len(columnas) + 1):
        cell = ws.cell(row=fila_suma_row, column=col_num)
        cell.font = Font(bold=True, color='000000')
        cell.alignment = center
        cell.border = border
        if col_num == 1:
            cell.fill = PatternFill(start_color='FFD966', end_color='FFD966', fill_type='solid')

    # Guardar el archivo en memoria
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Preparar respuesta HTTP para descargar Excel
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    filename = f"Ranking_Articulos_{datetime.today().strftime('%Y%m%d')}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'

    return response


def productos_horizontal(request):
    return render(request, 'paginaweb/productos_horizontal.html')


def productos_vertical(request):
    return render(request, 'paginaweb/productos_vertical.html')


def proyectos_destacados(request):
    return render(request, 'paginaweb/proyectos_destacados.html')


def manuales(request):
    return render(request, 'paginaweb/manuales.html')